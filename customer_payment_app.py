"""
Customer Payment Manager
A simple interface to look up customers and record payments
"""
 
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import calendar
import logging
import os
import fcntl
import shutil
import uuid
 
# === LOGGING ===
logging.basicConfig(
    filename='payment_app.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
 
# === COLUMN MAPPING ===
# All column references use names, resolved dynamically per sheet
COL = {
    'charge_date': 'Charge Date',
    'service': 'Service',
    'plan_cost': 'Plan Cost',
    'customer_name': 'Customer Name',
    'card_number': 'Card Number',
    'exp': 'Exp',
    'cvv': 'CVV',
    'amount_due': 'Amount Due',
    'status': 'Status',
    'phone': 'Phone',
    'notes': 'Notes',
    'due_day': 'Due Day',
    'notes2': 'Notes2',
    'square_customer_id': 'Square Customer ID',
    'square_card_id': 'Square Card ID',
}

# === SQUARE PAYMENT INTEGRATION ===
SQUARE_ACCESS_TOKEN = "EAAAlw_J8kMvMUP4nm3FPaBbrRTsgi5aM1VJb9UHgeA3cmdmqKnUguJnPmNtI40e"
SQUARE_LOCATION_ID  = "LJQJN0SC79G1M"

def _square_headers():
    """Return HTTP headers for Square API requests."""
    return {
        "Authorization": f"Bearer {SQUARE_ACCESS_TOKEN}",
        "Content-Type": "application/json",
        "Square-Version": "2024-01-18",
    }


def _square_errors(data):
    """Extract a readable error string from a Square API response dict."""
    errs = data.get("errors", [])
    if errs:
        return "; ".join(f"{e.get('code','')}: {e.get('detail','')}" for e in errs)
    return "Unknown error"


def save_square_ids(sheet_name, customer_name, square_customer_id, square_card_id):
    """Persist Square Customer ID and Card ID to Excel (adds columns if needed)."""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        headers = get_header_map(ws)
        for col_name in ['Square Customer ID', 'Square Card ID']:
            if col_name not in headers:
                max_col = ws.max_column + 1
                ws.cell(row=1, column=max_col, value=col_name)
                headers[col_name] = max_col
        for row in ws.iter_rows(min_row=2):
            if get_cell(row, headers, 'Customer Name') == customer_name:
                if square_customer_id is not None:
                    set_cell(row, headers, 'Square Customer ID', square_customer_id)
                if square_card_id is not None:
                    set_cell(row, headers, 'Square Card ID', square_card_id)
                break
        locked_save(wb, EXCEL_FILE)
        return True
    except Exception as e:
        logger.error(f"Error saving Square IDs for {customer_name}: {e}")
        return False


def create_square_customer(customer_name, phone):
    """Create a Square customer via REST API. Returns (customer_id or None, error or None)."""
    import requests as _req
    name_parts = str(customer_name).strip().split(' ', 1)
    body = {
        "idempotency_key": str(uuid.uuid4()),
        "given_name": name_parts[0],
        "phone_number": str(phone).strip() if phone else '',
    }
    if len(name_parts) > 1:
        body["family_name"] = name_parts[1]
    try:
        r = _req.post(
            "https://connect.squareup.com/v2/customers",
            json=body, headers=_square_headers(), timeout=15
        )
        data = r.json()
        if r.status_code == 200 and "customer" in data:
            return data["customer"]["id"], None
        return None, _square_errors(data)
    except Exception as e:
        return None, str(e)


def list_square_cards(square_customer_id):
    """Return (list of card dicts, error or None) for a Square customer via REST API."""
    import requests as _req
    try:
        r = _req.get(
            "https://connect.squareup.com/v2/cards",
            params={"customer_id": square_customer_id},
            headers=_square_headers(), timeout=15
        )
        data = r.json()
        if r.status_code == 200:
            return data.get("cards", []), None
        return [], _square_errors(data)
    except Exception as e:
        return [], str(e)


def square_charge_card(customer_name, amount, square_card_id, square_customer_id=None, note=""):
    """Charge a card on file via Square REST API. Returns (success, message, payment_id or None)."""
    import requests as _req
    if not square_card_id:
        return False, "No Square Card ID provided.", None
    body = {
        "source_id": square_card_id,
        "idempotency_key": str(uuid.uuid4()),
        "amount_money": {
            "amount": int(round(amount * 100)),
            "currency": "USD",
        },
        "location_id": SQUARE_LOCATION_ID,
        "note": note or f"K-Wireless payment for {customer_name}",
        "autocomplete": True,
    }
    if square_customer_id:
        body["customer_id"] = square_customer_id
    try:
        r = _req.post(
            "https://connect.squareup.com/v2/payments",
            json=body, headers=_square_headers(), timeout=15
        )
        data = r.json()
        if r.status_code == 200 and "payment" in data:
            pmt  = data["payment"]
            pid  = pmt.get("id", "")
            paid = pmt.get("amount_money", {}).get("amount", 0) / 100
            return True, f"Charged ${paid:.2f} via Square (ID: {pid})", pid
        return False, f"Card declined: {_square_errors(data)}", None
    except Exception as e:
        return False, f"Square API error: {str(e)}", None


# === 12-MONTH AGING SYSTEM ===
MONTHS_2026 = [
    'Jan_2026', 'Feb_2026', 'Mar_2026', 'Apr_2026', 'May_2026', 'Jun_2026',
    'Jul_2026', 'Aug_2026', 'Sep_2026', 'Oct_2026', 'Nov_2026', 'Dec_2026'
]
 
MONTH_MAP = {
    1: 'Jan_2026', 2: 'Feb_2026', 3: 'Mar_2026', 4: 'Apr_2026',
    5: 'May_2026', 6: 'Jun_2026', 7: 'Jul_2026', 8: 'Aug_2026',
    9: 'Sep_2026', 10: 'Oct_2026', 11: 'Nov_2026', 12: 'Dec_2026'
}
 
MONTHS_DISPLAY = [
    ('Jan_2026', 'January'), ('Feb_2026', 'February'), ('Mar_2026', 'March'),
    ('Apr_2026', 'April'), ('May_2026', 'May'), ('Jun_2026', 'June'),
    ('Jul_2026', 'July'), ('Aug_2026', 'August'), ('Sep_2026', 'September'),
    ('Oct_2026', 'October'), ('Nov_2026', 'November'), ('Dec_2026', 'December')
]
 
 
# === FILE CONFIGURATION ===
EXCEL_FILE_LOCAL = "cleaned_billing_by_service.xlsx"
RENDER_DISK_PATH = "/app/data"
RENDER_SRC_PATH = "/opt/render/project/src"
 
def resolve_excel_path():
    """Determine the correct Excel file path based on environment."""
    if os.path.exists(RENDER_DISK_PATH):
        disk_file = os.path.join(RENDER_DISK_PATH, "cleaned_billing_by_service.xlsx")
        if os.path.exists(disk_file):
            return disk_file
        src_file = os.path.join(RENDER_SRC_PATH, "cleaned_billing_by_service.xlsx")
        if os.path.exists(src_file):
            return src_file
    elif os.path.exists(RENDER_SRC_PATH):
        src_file = os.path.join(RENDER_SRC_PATH, "cleaned_billing_by_service.xlsx")
        if os.path.exists(src_file):
            return src_file
    return EXCEL_FILE_LOCAL
 
EXCEL_FILE = resolve_excel_path()
 
st.set_page_config(page_title="K-Wireless Payment Manager", page_icon="📡", layout="wide")
 
 
# === HELPER FUNCTIONS ===
 
def mask_card(card_number):
    """Return full card number."""
    card_str = str(card_number).strip() if card_number else ''
    if not card_str or card_str.lower() in ('nan', 'none', ''):
        return 'N/A'
    return card_str
 
 
def safe_float(val, default=0.0):
    """Safely convert a value to float."""
    if val is None:
        return default
    try:
        s = str(val).strip()
        if not s or s.lower() == 'nan':
            return default
        return float(s)
    except (ValueError, TypeError):
        return default
 
 
def get_header_map(ws):
    """Build a column name -> index mapping from the first row of a worksheet."""
    return {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value is not None}
 
 
def get_cell(row, headers, col_name, default=None):
    """Get a cell value from a row by column name."""
    col_idx = headers.get(col_name)
    if col_idx is None:
        return default
    val = row[col_idx - 1].value
    return val if val is not None else default
 
 
def set_cell(row, headers, col_name, value):
    """Set a cell value in a row by column name."""
    col_idx = headers.get(col_name)
    if col_idx is not None:
        row[col_idx - 1].value = value
        return True
    return False
 
 
def locked_save(wb, filepath):
    """Save workbook with file locking to prevent concurrent write corruption."""
    lock_path = filepath + '.lock'
    lock_fd = None
    try:
        try:
            lock_fd = open(lock_path, 'w')
            fcntl.flock(lock_fd, fcntl.LOCK_EX)
        except Exception:
            pass  # Lock unavailable — proceed anyway
        # Attempt backup but never let it block the save
        try:
            if os.path.exists(filepath):
                backup = filepath + '.backup'
                shutil.copy2(filepath, backup)
                os.chmod(backup, 0o644)
        except Exception as e:
            logger.warning(f"Backup skipped: {e}")
        wb.save(filepath)
        logger.info(f"Saved {filepath}")
    except Exception as e:
        logger.error(f"Error saving {filepath}: {e}")
        raise
    finally:
        try:
            if lock_fd:
                fcntl.flock(lock_fd, fcntl.LOCK_UN)
                lock_fd.close()
        except Exception:
            pass
 
 
# === MONTHLY BALANCE FUNCTIONS ===
 
def get_monthly_balances(customer):
    """Get monthly balance values from a customer dict."""
    balances = {}
    for month in MONTHS_2026:
        balances[month] = safe_float(customer.get(month, 0))
    return balances
 
 
def get_total_balance_from_months(balances):
    """Sum all monthly balances."""
    return sum(balances.values())
 
 
def update_amount_due_from_months(sheet_name, customer_name):
    """Update the Amount Due column with SUM formula for all 12 month columns."""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        headers = get_header_map(ws)
 
        for row in ws.iter_rows(min_row=2):
            name_val = get_cell(row, headers, 'Customer Name')
            if name_val == customer_name:
                row_num = row[0].row
                # Write Excel SUM formula so the file stays formula-based
                set_cell(row, headers, 'Amount Due', f'=SUM(N{row_num}:Y{row_num})')
                break
 
        locked_save(wb, EXCEL_FILE)
        return True
    except Exception as e:
        logger.error(f"Error updating amount due: {e}")
        return False
 
 
def save_monthly_balance(sheet_name, customer_name, month_label, amount):
    """Save a specific month's balance for a customer."""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        headers = get_header_map(ws)
 
        if month_label not in headers:
            logger.error(f"Month column {month_label} not found in {sheet_name}")
            return False
 
        for row in ws.iter_rows(min_row=2):
            if get_cell(row, headers, 'Customer Name') == customer_name:
                set_cell(row, headers, month_label, amount)
                break
 
        locked_save(wb, EXCEL_FILE)
        return True
    except Exception as e:
        logger.error(f"Error saving monthly balance: {e}")
        return False
 
 
# === CUSTOMER DATA FUNCTIONS ===
 
def save_customer_notes(sheet_name, customer_name, notes):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        headers = get_header_map(ws)
        for row in ws.iter_rows(min_row=2):
            if get_cell(row, headers, 'Customer Name') == customer_name:
                set_cell(row, headers, 'Notes', notes)
                break
        locked_save(wb, EXCEL_FILE)
        return True
    except Exception as e:
        logger.error(f"Error saving notes: {e}")
        st.error(f"Error saving notes: {e}")
        return False
 
 
def save_notes2(sheet_name, customer_name, notes2):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        headers = get_header_map(ws)
        for row in ws.iter_rows(min_row=2):
            if get_cell(row, headers, 'Customer Name') == customer_name:
                set_cell(row, headers, 'Notes2', notes2)
                break
        locked_save(wb, EXCEL_FILE)
        return True
    except Exception as e:
        logger.error(f"Error saving notes2: {e}")
        st.error(f"Error saving notes2: {e}")
        return False
 
 
def save_modem_number(sheet_name, customer_name, modem_number):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        headers = get_header_map(ws)
        for row in ws.iter_rows(min_row=2):
            if get_cell(row, headers, 'Customer Name') == customer_name:
                set_cell(row, headers, 'Modem Numbers', modem_number)
                break
        locked_save(wb, EXCEL_FILE)
        return True
    except Exception as e:
        logger.error(f"Error saving modem number: {e}")
        return False
 
 
def save_due_date(sheet_name, customer_name, due_day):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        headers = get_header_map(ws)
        for row in ws.iter_rows(min_row=2):
            if get_cell(row, headers, 'Customer Name') == customer_name:
                set_cell(row, headers, 'Due Day', due_day)
                break
        locked_save(wb, EXCEL_FILE)
        return True
    except Exception as e:
        logger.error(f"Error saving due date: {e}")
        st.error(f"Error saving due date: {e}")
        return False
 
 
def save_customer_info(sheet_name, customer_name, new_name, phone, card_number, exp, cvv, plan_cost, original_phone=None, row_index=None):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        headers = get_header_map(ws)
        target_row = None
 
        if row_index is not None:
            excel_row_num = row_index + 2
            for row in ws.iter_rows(min_row=excel_row_num, max_row=excel_row_num):
                target_row = row
                break
        else:
            for row in ws.iter_rows(min_row=2):
                if str(get_cell(row, headers, 'Customer Name', '')).strip() == str(customer_name).strip():
                    target_row = row
                    break
 
        if target_row is None:
            return False
 
        set_cell(target_row, headers, 'Plan Cost', plan_cost)
        set_cell(target_row, headers, 'Customer Name', new_name)
        set_cell(target_row, headers, 'Card Number', card_number)
        set_cell(target_row, headers, 'Exp', exp)
        set_cell(target_row, headers, 'CVV', cvv)
        set_cell(target_row, headers, 'Phone', phone)
 
        locked_save(wb, EXCEL_FILE)
        logger.info(f"Updated customer info for {customer_name} in {sheet_name}")
        return True
    except Exception as e:
        logger.error(f"Error saving customer info: {e}")
        st.error(f"Error saving customer info: {e}")
        return False
 
 
def get_balance(customer):
    """Calculate customer balance: prefer sum of monthly columns, fallback to Amount Due."""
    # First try monthly totals
    total_monthly = 0
    has_monthly = False
    for month in MONTHS_2026:
        val = safe_float(customer.get(month, 0))
        if val != 0:
            has_monthly = True
        total_monthly += val
 
    if has_monthly:
        return total_monthly
 
    # Fallback to Amount Due column
    return safe_float(customer.get('Amount Due', 0))
 
 
# === PAYMENT FUNCTIONS ===
 
def save_payment(sheet_name, customer_name, payment_amount, pay_month=None, notes="", advance_due=False):
    """Save a payment. If pay_month specified, deduct from that month's balance.
 
    If advance_due=True, also advances the Due Day by 30 days from the current
    due date — all in one atomic save to the Excel file.
    Returns (success: bool, new_due_date: datetime or None).
    """
    new_due_date = None
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        headers = get_header_map(ws)
 
        for row in ws.iter_rows(min_row=2):
            if get_cell(row, headers, 'Customer Name') != customer_name:
                continue
 
            now = datetime.now()
            timestamp = now.strftime('%Y-%m-%d %I:%M %p')
 
            if pay_month and pay_month in headers:
                # Apply payment to specific month — clear the balance
                current_val = safe_float(get_cell(row, headers, pay_month, 0))
                new_val = max(0, current_val - payment_amount)
                set_cell(row, headers, pay_month, new_val)
            else:
                # Apply to Amount Due directly
                current_balance = safe_float(get_cell(row, headers, 'Amount Due', 0))
                new_balance = current_balance - payment_amount
                set_cell(row, headers, 'Amount Due', new_balance)
 
            # Recalculate total Amount Due from all months
            total = 0
            for m in MONTHS_2026:
                total += safe_float(get_cell(row, headers, m, 0))
            row_num = row[0].row
            set_cell(row, headers, 'Amount Due', f'=SUM(N{row_num}:Y{row_num})')
 
            # Update status
            if total <= 0:
                set_cell(row, headers, 'Status', 'Paid')
            else:
                set_cell(row, headers, 'Status', 'Partial')
 
            # Advance due date if requested (current month payment only)
            # Rule: always advance 30 days from Charge Date (the original due date),
            # regardless of when the payment is actually made.
            # If auto-charge already moved Charge Date to the future, leave it alone.
            if advance_due:
                try:
                    existing_charge_date = get_cell(row, headers, 'Charge Date')
                    base_date = None
                    if existing_charge_date:
                        try:
                            if isinstance(existing_charge_date, datetime):
                                base_date = existing_charge_date
                            else:
                                base_date = datetime.strptime(str(existing_charge_date).strip()[:10], '%Y-%m-%d')
                        except Exception:
                            pass
 
                    if base_date and base_date.date() > now.date():
                        # Auto-charge already advanced the date — keep it, just record it in notes
                        new_due_date = base_date
                    else:
                        # Advance exactly 30 days from the Charge Date (due date), not from today
                        if base_date is None:
                            # Fallback: use Due Day in current month
                            current_due = get_cell(row, headers, 'Due Day')
                            current_due_int = int(current_due) if current_due is not None else now.day
                            max_day = calendar.monthrange(now.year, now.month)[1]
                            base_date = datetime(now.year, now.month, min(current_due_int, max_day))
                        new_due_date = base_date + timedelta(days=30)
                        set_cell(row, headers, 'Due Day', new_due_date.day)
                        set_cell(row, headers, 'Charge Date', new_due_date.strftime('%Y-%m-%d'))
                        logger.info(f"Due date for {customer_name}: {base_date.strftime('%m/%d/%Y')} -> {new_due_date.strftime('%m/%d/%Y')}")
                except (ValueError, TypeError) as e:
                    logger.error(f"Error advancing due date for {customer_name}: {e}")
 
            # Append payment with date AND time to Notes
            existing_notes = str(get_cell(row, headers, 'Notes', '') or '')
            month_label = dict(MONTHS_DISPLAY).get(pay_month, pay_month) if pay_month else ''
            payment_info = f"Paid ${payment_amount:.2f}"
            if month_label:
                payment_info += f" ({month_label})"
            payment_info += f" on {timestamp}"
            if new_due_date:
                payment_info += f" | Next due: {new_due_date.strftime('%m/%d/%Y')}"
            if notes:
                payment_info += f" - {notes}"
            new_notes = (existing_notes + " | " + payment_info) if existing_notes else payment_info
            set_cell(row, headers, 'Notes', new_notes)
 
            break
 
        locked_save(wb, EXCEL_FILE)
        logger.info(f"Payment ${payment_amount:.2f} recorded for {customer_name} in {sheet_name}")
        return True, new_due_date
    except Exception as e:
        logger.error(f"Error saving payment: {e}")
        st.error(f"Error saving: {e}")
        return False, None


def parse_payment_notes(notes_str, customer_name, sheet_name):
    """Parse Notes field to extract individual payment entries.
    Returns list of dicts: {customer, sheet, amount, month, method, dt}
    """
    import re as _re
    entries = []
    if not notes_str:
        return entries
    parts = str(notes_str).split(' | ')
    for part in parts:
        part = part.strip()
        # Match: "Paid $X.XX (Month Label) on YYYY-MM-DD HH:MM AM/PM"
        m = _re.match(
            r'Paid \$([0-9.]+)(?:\s+\(([^)]+)\))?\s+on\s+(\d{4}-\d{2}-\d{2}\s+\d{1,2}:\d{2}\s+[AP]M)',
            part
        )
        if not m:
            continue
        try:
            amount = float(m.group(1))
            month  = m.group(2) or ''
            dt_str = m.group(3)
            dt = datetime.strptime(dt_str, '%Y-%m-%d %I:%M %p')
        except Exception:
            continue
        method = 'Square' if ('Square:' in part or 'Square ID:' in part) else 'Manual'
        entries.append({
            'customer': customer_name,
            'sheet':    sheet_name,
            'amount':   amount,
            'month':    month,
            'method':   method,
            'dt':       dt,
        })
    return entries


def get_collections_report(start_date, end_date):
    """Scan all sheets and return payment entries within the date range.
    Returns (list_of_entries, error_string_or_None).
    """
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
    except Exception as e:
        return [], str(e)
    all_entries = []
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            headers = get_header_map(ws)
            if 'Customer Name' not in headers or 'Notes' not in headers:
                continue
            for row in ws.iter_rows(min_row=2):
                cust_name = get_cell(row, headers, 'Customer Name')
                notes     = get_cell(row, headers, 'Notes', '')
                if not cust_name or not notes:
                    continue
                row_entries = parse_payment_notes(str(notes), str(cust_name), sheet_name)
                for e in row_entries:
                    if start_date <= e['dt'].date() <= end_date:
                        all_entries.append(e)
        except Exception:
            continue
    wb.close()
    all_entries.sort(key=lambda x: x['dt'], reverse=True)
    return all_entries, None


def _build_report_html(entries, period_label, total, sq_total, manual_total):
    """Build a printable HTML collections report."""
    rows_html = ''
    for e in entries:
        if e['method'] == 'Square':
            badge = '<span style="background:#28a745;color:white;padding:2px 8px;border-radius:10px;font-size:11px">Square</span>'
        else:
            badge = '<span style="background:#6c757d;color:white;padding:2px 8px;border-radius:10px;font-size:11px">Manual</span>'
        rows_html += (
            f'<tr>'
            f'<td>{e["customer"]}</td>'
            f'<td>{e["sheet"]}</td>'
            f'<td style="text-align:right;font-weight:bold">${e["amount"]:.2f}</td>'
            f'<td>{e["month"]}</td>'
            f'<td>{badge}</td>'
            f'<td>{e["dt"].strftime("%m/%d/%Y %I:%M %p")}</td>'
            f'</tr>'
        )
    generated = datetime.now().strftime('%m/%d/%Y %I:%M %p')
    return (
        '<!DOCTYPE html><html><head><meta charset="UTF-8">'
        '<title>K-Wireless Collections Report</title>'
        '<style>'
        'body{font-family:Arial,sans-serif;margin:40px;color:#333}'
        'h1{color:#1a1a2e;font-size:22px;margin-bottom:4px}'
        '.subtitle{color:#666;font-size:13px;margin-bottom:20px}'
        'table{width:100%;border-collapse:collapse;margin-top:15px;font-size:13px}'
        'th{background:#1a1a2e;color:white;padding:10px;text-align:left}'
        'td{padding:8px 10px;border-bottom:1px solid #eee}'
        'tr:nth-child(even){background:#f8f9fa}'
        '.summary{margin-top:24px;padding:15px 20px;background:#f0f4ff;border-radius:8px;border-left:4px solid #1a1a2e}'
        '.total{font-size:18px;font-weight:bold;color:#1a1a2e;margin:0 0 6px 0}'
        '.detail{font-size:13px;color:#555;margin:3px 0}'
        'button{padding:10px 20px;background:#1a1a2e;color:white;border:none;border-radius:6px;cursor:pointer;font-size:14px;margin-top:16px}'
        '@media print{button{display:none}}'
        '</style></head><body>'
        f'<h1>&#128200; K-Wireless Collections Report</h1>'
        f'<div class="subtitle">Period: {period_label} &nbsp;|&nbsp; Generated: {generated}</div>'
        '<table><thead><tr>'
        '<th>Customer</th><th>Service</th><th style="text-align:right">Amount</th>'
        '<th>Month</th><th>Method</th><th>Date / Time</th>'
        f'</tr></thead><tbody>{rows_html}</tbody></table>'
        '<div class="summary">'
        f'<p class="total">Total Collected: ${total:.2f}</p>'
        f'<p class="detail">&#128179; Square: ${sq_total:.2f} &nbsp;&nbsp; &#9995; Manual: ${manual_total:.2f}</p>'
        f'<p class="detail">Transactions: {len(entries)}</p>'
        '</div>'
        '<button onclick="window.print()">&#128438; Print Report</button>'
        '</body></html>'
    )


def get_last_payment_with_month(notes_str):
    """Parse the most recent payment entry from a Notes string including month label."""
    import re
    pattern = re.compile(
        r'Paid \$([0-9,.]+)(?:\s+\(([^)]+)\))?\s+on\s+(\d{4}-\d{2}-\d{2}[^|]*)(?:\|\s*Next due:\s*(\d{2}/\d{2}/\d{4}))?'
    )
    matches = list(pattern.finditer(notes_str))
    if not matches:
        return None
    m = matches[-1]
    return {
        'amount':     float(m.group(1).replace(',', '')),
        'month_col':  m.group(2),
        'paid_on':    m.group(3).strip(),
        'next_due':   m.group(4),
        'full_entry': m.group(0),
    }


def rollback_payment(sheet_name, customer_name):
    """Reverse the most recent payment for a customer.
    Restores month balance, removes payment note, and reverses due date advance."""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        headers = get_header_map(ws)

        for row in ws.iter_rows(min_row=2):
            if get_cell(row, headers, 'Customer Name') != customer_name:
                continue

            notes = str(get_cell(row, headers, 'Notes', '') or '')
            last = get_last_payment_with_month(notes)
            if not last:
                return False, "No payment found in notes to roll back."

            # Restore month column balance
            month_col = last['month_col']
            if month_col and month_col in headers:
                set_cell(row, headers, month_col, last['amount'])

            # Reverse due date advance: subtract 30 days from Next due
            original_due = None
            if last['next_due']:
                try:
                    next_due_dt = datetime.strptime(last['next_due'], '%m/%d/%Y')
                    original_due = next_due_dt - timedelta(days=30)
                    set_cell(row, headers, 'Due Day', original_due.day)
                    set_cell(row, headers, 'Charge Date', original_due.strftime('%Y-%m-%d'))
                except Exception:
                    pass

            # Remove the last payment entry from Notes
            entry = last['full_entry']
            if f" | {entry}" in notes:
                new_notes = notes.replace(f" | {entry}", "")
            elif notes.strip() == entry.strip():
                new_notes = ''
            else:
                new_notes = notes.replace(entry, "").strip().strip('|').strip()
            set_cell(row, headers, 'Notes', new_notes)

            # Recalculate Amount Due
            row_num = row[0].row
            set_cell(row, headers, 'Amount Due', f'=SUM(N{row_num}:Y{row_num})')

            # Clear status
            set_cell(row, headers, 'Status', '')

            break

        locked_save(wb, EXCEL_FILE)
        msg = f"Rolled back ${last['amount']:.2f}"
        if last['month_col']:
            msg += f" ({last['month_col']})"
        if original_due:
            msg += f". Due date restored to {original_due.strftime('%m/%d/%Y')}."
        logger.info(f"Rollback for {customer_name} ({sheet_name}): {msg}")
        return True, msg
    except Exception as e:
        logger.error(f"Rollback error for {customer_name}: {e}")
        return False, str(e)


def advance_due_date(sheet_name, customer_name, days=30):
    """Advance the due date by calculating 30 days from the current due date.
 
    Uses the current due day + current month/year to build an actual date,
    adds 30 days, and stores the new day-of-month as the Due Day.
    Also updates the Charge Date to the new due date.
    """
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[sheet_name]
        headers = get_header_map(ws)
 
        new_due_date = None
        for row in ws.iter_rows(min_row=2):
            if get_cell(row, headers, 'Customer Name') == customer_name:
                current_due = get_cell(row, headers, 'Due Day')
                if current_due is not None:
                    try:
                        current_due_int = int(current_due)
                        now = datetime.now()
                        # Build the actual due date from the due day + current month/year
                        # Clamp the day to the max days in the current month
                        max_day = calendar.monthrange(now.year, now.month)[1]
                        due_day_clamped = min(current_due_int, max_day)
                        current_due_date = datetime(now.year, now.month, due_day_clamped)
                        # Add 30 days to get next due date
                        new_due_date = current_due_date + timedelta(days=days)
                        set_cell(row, headers, 'Due Day', new_due_date.day)
                        # Update Charge Date to new due date
                        set_cell(row, headers, 'Charge Date', new_due_date.strftime('%Y-%m-%d'))
                        logger.info(f"Due date for {customer_name}: {current_due_date.strftime('%m/%d')} -> {new_due_date.strftime('%m/%d/%Y')}")
                    except (ValueError, TypeError) as e:
                        logger.error(f"Error calculating due date for {customer_name}: {e}")
                break
 
        locked_save(wb, EXCEL_FILE)
        return new_due_date
    except Exception as e:
        logger.error(f"Error advancing due date: {e}")
        return None
 
 
# === DATA LOADING & SEARCH ===
 
def load_excel():
    """Load all sheets from the Excel file into DataFrames.
 
    Note: Amount Due may contain Excel formulas (=SUM(N:Y)).
    pandas/openpyxl with data_only may return None for formulas if the file
    hasn't been opened in Excel. We recalculate Amount Due from monthly columns.
    """
    try:
        excel_file = pd.ExcelFile(EXCEL_FILE)
        all_data = {}
        for sheet in excel_file.sheet_names:
            if sheet in ("Summary", "1"):
                continue
            df = pd.read_excel(excel_file, sheet_name=sheet)
            if df.empty:
                continue
            df['Service'] = sheet
 
            # Ensure Due Day column exists
            if 'Due Day' not in df.columns:
                if 'Charge Date' in df.columns:
                    try:
                        df['Charge Date'] = pd.to_datetime(df['Charge Date'], errors='coerce')
                        df['Due Day'] = df['Charge Date'].dt.day
                    except Exception:
                        df['Due Day'] = None
 
            # Recalculate Amount Due from monthly columns
            # (handles formula strings and None values from openpyxl)
            for idx, row in df.iterrows():
                monthly_total = 0
                for m in MONTHS_2026:
                    monthly_total += safe_float(row.get(m, 0))
                if monthly_total > 0:
                    df.at[idx, 'Amount Due'] = monthly_total
                else:
                    # Try to use existing Amount Due if it's a real number
                    existing = row.get('Amount Due', 0)
                    df.at[idx, 'Amount Due'] = safe_float(existing)
 
            all_data[sheet] = df
        return all_data, excel_file
    except Exception as e:
        logger.error(f"Error loading file: {e}")
        st.error(f"Error loading file: {e}")
        return None, None
 
 
def parse_collections_from_notes(all_data):
    """Parse Notes column across all sheets to extract payment records by date."""
    import re
    records = []
    pattern = re.compile(r'Paid \$([0-9,.]+)(?:\s+\([^)]+\))?\s+on\s+(\d{4}-\d{2}-\d{2})')
    for service, df in (all_data or {}).items():
        if df is None or df.empty:
            continue
        for _, row in df.iterrows():
            notes = str(row.get('Notes', '') or '')
            customer = row.get('Customer Name', 'Unknown')
            for match in pattern.finditer(notes):
                try:
                    amount = float(match.group(1).replace(',', ''))
                    date = datetime.strptime(match.group(2), '%Y-%m-%d').date()
                    records.append({
                        'date': date,
                        'customer': customer,
                        'service': service,
                        'amount': amount
                    })
                except Exception:
                    continue
    return records
 
 
def _build_customer_result(row, service, idx=None):
    """Build a standardized customer result dict from a DataFrame row."""
    monthly = {}
    for m in MONTHS_2026:
        monthly[m] = row.get(m, 0)
 
    result = {
        'Service': service,
        'Customer Name': row.get('Customer Name', ''),
        'Phone': row.get('Phone', ''),
        'Amount Due': row.get('Amount Due', ''),
        'Status': row.get('Status', ''),
        'Card Number': row.get('Card Number', ''),
        'Exp': row.get('Exp', ''),
        'CVV': row.get('CVV', ''),
        'Plan Cost': row.get('Plan Cost', ''),
        'Charge Date': row.get('Charge Date', ''),
        'Due Day': row.get('Due Day', ''),
        'Notes': row.get('Notes', ''),
        'Notes2': row.get('Notes2', ''),
        'Payment Date': row.get('Payment Date', ''),
        'Modem Numbers': row.get('Modem Numbers', ''),
        'Square Customer ID': row.get('Square Customer ID', ''),
        'Square Card ID': row.get('Square Card ID', ''),
    }
    if idx is not None:
        result['row_index'] = idx
    result.update(monthly)
    return result
 
 
def search_customers(all_data, query):
    """Search customers by name, phone, or card number."""
    results = []
    # Escape special regex characters in the query
    import re
    escaped_query = re.escape(query)
    for service, df in all_data.items():
        if df is None or df.empty:
            continue
        mask = (
            df['Customer Name'].astype(str).str.contains(escaped_query, case=False, na=False) |
            df['Phone'].astype(str).str.contains(escaped_query, case=False, na=False) |
            df['Card Number'].astype(str).str.contains(escaped_query, case=False, na=False)
        )
        matches = df[mask]
        for idx, row in matches.iterrows():
            results.append(_build_customer_result(row, service, idx))
    return results
 
 
def get_customers_by_due_day(all_data, due_day):
    results = []
    for service, df in all_data.items():
        if df is None or df.empty:
            continue
        if 'Due Day' in df.columns:
            # Compare as integers to handle mixed types
            for _, row in df.iterrows():
                try:
                    row_due = int(row.get('Due Day', 0)) if row.get('Due Day') is not None and not pd.isna(row.get('Due Day')) else None
                except (ValueError, TypeError):
                    row_due = None
                if row_due == due_day:
                    results.append(_build_customer_result(row, service))
    return results
 
 
def auto_charge_due_today():
    """Auto-charge customers due today OR missed in the past 7 days.
    Skips anyone whose Charge Date is already in the future (paid up) or already charged this month."""
    today = datetime.now()
    current_month = today.month
    current_year = today.year
    month_key = MONTH_MAP.get(current_month, 'Mar_2026')
 
    # Build list of day numbers to catch up on (today + up to 6 days back, current month only)
    days_to_check = []
    for i in range(7):
        check_date = today - timedelta(days=i)
        if check_date.month == current_month:
            days_to_check.append(check_date.day)
 
    charged = []
    try:
        wb = load_workbook(EXCEL_FILE)
 
        for sheet_name in wb.sheetnames:
            if sheet_name == 'Summary':
                continue
            ws = wb[sheet_name]
            headers = get_header_map(ws)
 
            # Ensure month columns exist
            if month_key not in headers:
                max_col = ws.max_column
                for m in MONTHS_2026:
                    if m not in headers:
                        max_col += 1
                        ws.cell(row=1, column=max_col, value=m)
                headers = get_header_map(ws)
 
            if not all(headers.get(k) for k in ['Due Day', 'Plan Cost', month_key]):
                continue
 
            for row in ws.iter_rows(min_row=2):
                due_day = get_cell(row, headers, 'Due Day')
                plan_cost = get_cell(row, headers, 'Plan Cost')
                customer_name = get_cell(row, headers, 'Customer Name')
 
                try:
                    due_day_int = int(due_day) if due_day is not None else None
                except (ValueError, TypeError):
                    due_day_int = None
 
                # Use Charge Date as the authoritative due date for window and skip checks
                charge_date_val = get_cell(row, headers, 'Charge Date')
                charge_date_parsed = None
                if charge_date_val:
                    try:
                        if isinstance(charge_date_val, datetime):
                            charge_date_parsed = charge_date_val.date()
                        else:
                            charge_date_parsed = datetime.strptime(str(charge_date_val).strip()[:10], '%Y-%m-%d').date()
                    except Exception:
                        pass
 
                # Skip if Charge Date is in the future — due date already advanced, customer is paid up
                if charge_date_parsed and charge_date_parsed > today.date():
                    continue
 
                # Only process if Charge Date falls within our 7-day catchup window
                # Fall back to Due Day number check if no Charge Date
                if charge_date_parsed:
                    cutoff = today.date() - timedelta(days=6)
                    if not (cutoff <= charge_date_parsed <= today.date()):
                        continue
                else:
                    if due_day_int not in days_to_check:
                        continue
 
                # Skip if already charged this month but not yet paid (month column has a value)
                current_month_val = safe_float(get_cell(row, headers, month_key, 0))
                if current_month_val > 0:
                    continue
 
                try:
                    charge_amount = safe_float(plan_cost)
                    if charge_amount <= 0:
                        continue

                    # === Square charging (if customer has a card on file) ===
                    sq_card_raw = get_cell(row, headers, 'Square Card ID')
                    sq_cust_raw = get_cell(row, headers, 'Square Customer ID')
                    sq_card_id  = str(sq_card_raw).strip() if sq_card_raw else ''
                    sq_cust_id  = str(sq_cust_raw).strip() if sq_cust_raw else ''
                    if sq_card_id.lower()  in ('nan', 'none', ''): sq_card_id  = ''
                    if sq_cust_id.lower()  in ('nan', 'none', ''): sq_cust_id  = ''

                    square_payment_id = None
                    square_error      = None
                    if sq_card_id:
                        sq_ok, sq_msg, sq_pid = square_charge_card(
                            customer_name, charge_amount, sq_card_id,
                            sq_cust_id or None,
                            note=f"K-Wireless auto-charge {today.strftime('%m/%d/%Y')}"
                        )
                        if sq_ok:
                            square_payment_id = sq_pid
                        else:
                            square_error = sq_msg

                    # Set month balance: 0 (paid) if Square succeeded, else charge_amount (owed)
                    if square_payment_id:
                        set_cell(row, headers, month_key, 0)
                    else:
                        set_cell(row, headers, month_key, charge_amount)

                    # Update total Amount Due with formula
                    row_num = row[0].row
                    set_cell(row, headers, 'Amount Due', f'=SUM(N{row_num}:Y{row_num})')

                    # Advance due date 30 days from Charge Date (the authoritative due date)
                    new_due_date = None
                    try:
                        base_date = None
                        if charge_date_val:
                            try:
                                if isinstance(charge_date_val, datetime):
                                    base_date = charge_date_val
                                else:
                                    base_date = datetime.strptime(str(charge_date_val).strip()[:10], '%Y-%m-%d')
                            except Exception:
                                pass
                        if base_date is None and due_day_int is not None:
                            max_day = calendar.monthrange(current_year, current_month)[1]
                            due_day_clamped = min(due_day_int, max_day)
                            base_date = datetime(current_year, current_month, due_day_clamped)
                        if base_date:
                            new_due_date = base_date + timedelta(days=30)
                            set_cell(row, headers, 'Due Day', new_due_date.day)
                            set_cell(row, headers, 'Charge Date', new_due_date.strftime('%Y-%m-%d'))
                    except Exception:
                        pass

                    # Write note for Square charges so collections parser picks them up
                    if square_payment_id or square_error:
                        existing_notes = str(get_cell(row, headers, 'Notes', '') or '')
                        ts = today.strftime('%Y-%m-%d %I:%M %p')
                        if square_payment_id:
                            note_entry = (
                                f"Paid ${charge_amount:.2f} ({month_key}) on {ts}"
                                f" | Square: {square_payment_id}"
                            )
                            if new_due_date:
                                note_entry += f" | Next due: {new_due_date.strftime('%m/%d/%Y')}"
                        else:
                            short_err = (square_error or '')[:80]
                            note_entry = f"Square DECLINED ${charge_amount:.2f} on {today.strftime('%Y-%m-%d')}: {short_err}"
                        new_notes = (existing_notes + " | " + note_entry) if existing_notes else note_entry
                        set_cell(row, headers, 'Notes', new_notes)

                    # Record in results list
                    if square_payment_id:
                        charged.append(f"{customer_name} ({sheet_name}): ${charge_amount:.2f} ✅ Charged via Square")
                    elif square_error:
                        charged.append(f"{customer_name} ({sheet_name}): ${charge_amount:.2f} ❌ Square declined — balance added")
                    elif due_day_int is not None and due_day_int != today.day:
                        charged.append(f"{customer_name} ({sheet_name}): ${charge_amount:.2f} ⚠️ catch-up from day {due_day_int}")
                    else:
                        charged.append(f"{customer_name} ({sheet_name}): ${charge_amount:.2f}")

                    logger.info(
                        f"Auto-charged {customer_name} ({sheet_name}): ${charge_amount:.2f}"
                        + (f" via Square {square_payment_id}" if square_payment_id else "")
                    )
                except Exception as e:
                    logger.error(f"Error charging {customer_name}: {e}")
 
        locked_save(wb, EXCEL_FILE)
        return charged
    except Exception as e:
        logger.error(f"Auto-charge error: {e}")
        return [f"Error: {str(e)}"]
 
 
def get_past_due_customers(all_data):
    """Return customers who are genuinely past due:
    - Any balance in a PREVIOUS month (always past due), OR
    - A balance in the CURRENT month only if their Charge Date is today or in the past.
    Customers whose balance is only in the current month with a future Charge Date are excluded."""
    results = []
    today = datetime.now().date()
    current_month_key = MONTH_MAP.get(today.month)
    current_month_idx = MONTHS_2026.index(current_month_key) if current_month_key in MONTHS_2026 else -1
    past_month_keys = MONTHS_2026[:current_month_idx] if current_month_idx > 0 else []

    for service, df in all_data.items():
        if df is None or df.empty:
            continue
        for _, row in df.iterrows():
            status = row.get('Status', '')
            status_str = str(status).upper() if status else ''
            if 'PAID' in status_str or status_str == 'READY':
                continue

            # Check for balance in previous months — always past due
            has_prev_balance = any(safe_float(row.get(m, 0)) > 0 for m in past_month_keys)

            # Check for balance in current month
            current_month_balance = safe_float(row.get(current_month_key, 0)) if current_month_key else 0
            has_current_balance = current_month_balance > 0

            # For current month balance, only count as past due if Charge Date <= today
            current_month_due = False
            if has_current_balance:
                charge_date_raw = row.get('Charge Date')
                if charge_date_raw is not None and not pd.isna(charge_date_raw):
                    try:
                        if isinstance(charge_date_raw, datetime):
                            charge_date = charge_date_raw.date()
                        else:
                            charge_date = datetime.strptime(str(charge_date_raw).strip()[:10], '%Y-%m-%d').date()
                        current_month_due = (charge_date <= today)
                    except Exception:
                        current_month_due = True  # Can't parse date — include to be safe
                else:
                    current_month_due = True  # No charge date — include to be safe

            if has_prev_balance or current_month_due:
                results.append(_build_customer_result(row, service))
    return results
 
 
def get_collections_by_date(all_data, start_date, end_date):
    results = []
    for service, df in all_data.items():
        if df is None or df.empty:
            continue
        if 'Payment Date' not in df.columns:
            continue
        for _, row in df.iterrows():
            payment_date = row.get('Payment Date')
            if payment_date is None or pd.isna(payment_date):
                continue
            try:
                payment_date = pd.to_datetime(payment_date)
            except (ValueError, TypeError):
                continue
            if start_date <= payment_date.date() <= end_date:
                results.append({
                    'Service': service,
                    'Customer Name': row.get('Customer Name', ''),
                    'Phone': row.get('Phone', ''),
                    'Amount Collected': safe_float(row.get('Amount Due', 0)),
                    'Status': row.get('Status', ''),
                    'Payment Date': payment_date.strftime('%Y-%m-%d'),
                    'Notes': row.get('Notes', '')
                })
    return results
 
 
# === UI HELPERS ===
 
def display_customer_card(customer, index):
    """Display a single customer's details and action forms."""
    charge_date = customer.get('Charge Date', '')
    if charge_date:
        try:
            charge_date = pd.to_datetime(charge_date).strftime('%Y-%m-%d')
        except (ValueError, TypeError):
            pass
 
    balance = get_balance(customer)
    monthly_balances = get_monthly_balances(customer)
    total_from_aging = get_total_balance_from_months(monthly_balances)
 
    # Balance display
    if balance < 0:
        balance_color = "green"
        balance_display = f"CREDIT: ${abs(balance):.2f}"
    elif balance == 0:
        balance_color = "gray"
        balance_display = "$0.00"
    else:
        balance_color = "red"
        balance_display = f"${balance:.2f}"
 
    status = str(customer.get('Status', '') or '').strip()
    if not status:
        status = 'No Status'
 
    # Header
    st.markdown(f"### {customer['Customer Name']}")
 
    modem = str(customer.get('Modem Numbers', '') or '').strip()
    if not modem or modem.lower() in ('nan', 'none'):
        modem = ''
 
    col_info1, col_info2, col_info3 = st.columns(3)
    with col_info1:
        st.write(f"**Service:** {customer['Service']}")
        st.write(f"**Phone:** {customer.get('Phone', 'N/A')}")
    with col_info2:
        st.write(f"**Status:** {status}")
        st.write(f"**Due Day:** {customer.get('Due Day', 'N/A')}")
    with col_info3:
        st.write(f"**Plan Cost:** ${safe_float(customer.get('Plan Cost', 0)):.2f}")
        card_display = mask_card(customer.get('Card Number', ''))
        exp_display = customer.get('Exp', 'N/A')
        cvv_display = customer.get('CVV', 'N/A')
        st.write(f"**Card:** {card_display} | **Exp:** {exp_display} | **CVV:** {cvv_display}")
        sq_card_chk = str(customer.get('Square Card ID', '') or '').strip()
        if sq_card_chk and sq_card_chk.lower() not in ('nan', 'none', ''):
            st.write("**Square:** ✅ Card on file")
        else:
            st.write("**Square:** ⬜ No card linked")
    if modem:
        st.write(f"**Modem Number:** {modem}")
 
    st.markdown(f"**Total Balance: :{balance_color}[{balance_display}]**")
 

    # Pre-compute Square IDs used by multiple sections below
    sq_card_id = str(customer.get('Square Card ID', '')  or '').strip()
    sq_cust_id = str(customer.get('Square Customer ID', '') or '').strip()
    if sq_card_id.lower() in ('nan', 'none', ''): sq_card_id = ''
    if sq_cust_id.lower() in ('nan', 'none', ''): sq_cust_id = ''
    current_month_col = MONTH_MAP.get(datetime.now().month)


    # ── 1. Charge Card via Square ──────────────────────────
    if sq_card_id:
        with st.expander("💳 Charge Card via Square"):
            sq_pay_month = st.selectbox(
                "Apply to Month:", options=[m[0] for m in MONTHS_DISPLAY],
                format_func=lambda x: dict(MONTHS_DISPLAY).get(x, x),
                key=f"sq_pay_month_{index}"
            )
            month_bal = monthly_balances.get(sq_pay_month, 0)
            default_amt = float(month_bal) if month_bal > 0 else float(safe_float(customer.get('Plan Cost', 0)))
            sq_amount = st.number_input(
                "Charge Amount ($):", min_value=0.01, value=default_amt,
                step=5.0, key=f"sq_amt_{index}"
            )
            sq_note = st.text_input("Note (optional):", key=f"sq_note_{index}")

            current_month_col = MONTH_MAP.get(datetime.now().month)
            if sq_pay_month == current_month_col:
                st.caption("Current month — due date will advance 30 days after charge.")

            if st.button(f"💳 Charge ${sq_amount:.2f} via Square", key=f"sq_charge_{index}"):
                with st.spinner("Processing Square payment..."):
                    sq_ok, sq_msg, sq_pid = square_charge_card(
                        customer['Customer Name'], sq_amount, sq_card_id,
                        sq_cust_id or None,
                        note=sq_note or f"K-Wireless payment"
                    )
                if sq_ok:
                    # Record the payment in Excel
                    is_current = (sq_pay_month == current_month_col)
                    p_ok, new_due = save_payment(
                        customer['Service'], customer['Customer Name'],
                        sq_amount, pay_month=sq_pay_month,
                        notes=f"Square ID: {sq_pid}",
                        advance_due=is_current
                    )
                    if p_ok:
                        if is_current and new_due:
                            st.success(f"✅ {sq_msg} — Next due: {new_due.strftime('%m/%d/%Y')}")
                        else:
                            st.success(f"✅ {sq_msg}")
                        st.rerun()
                    else:
                        st.warning(f"Square charged successfully ({sq_pid}) but Excel update failed. Please record manually.")
                else:
                    st.error(f"❌ {sq_msg}")


    # ── 2. Rollback Last Payment ───────────────────────────
    notes_val = str(customer.get('Notes', '') or '')
    last_pmt = get_last_payment_with_month(notes_val)
    if last_pmt:
        with st.expander("↩️ Rollback Last Payment"):
            st.warning(
                f"**Last payment:** ${last_pmt['amount']:.2f}"
                + (f" ({last_pmt['month_col']})" if last_pmt['month_col'] else "")
                + f" on {last_pmt['paid_on']}"
                + (f" | Next due was: {last_pmt['next_due']}" if last_pmt['next_due'] else "")
            )
            st.caption("This will restore the month balance, remove the payment note, and reverse the due date advance.")
            if st.button("⚠️ Confirm Rollback", key=f"rollback_{index}"):
                success, msg = rollback_payment(customer['Service'], customer['Customer Name'])
                if success:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(f"Rollback failed: {msg}")


    # ── 3. Post Payment ────────────────────────────────────
    with st.expander("Post Payment"):
        col_pay1, col_pay2 = st.columns(2)
        with col_pay1:
            pay_month = st.selectbox("Apply to Month:", options=[m[0] for m in MONTHS_DISPLAY],
                                     format_func=lambda x: dict(MONTHS_DISPLAY).get(x, x),
                                     key=f"pay_month_{index}")
        with col_pay2:
            month_balance = monthly_balances.get(pay_month, 0)
            pay_amount = st.number_input("Payment Amount:", min_value=0.0,
                                         value=float(month_balance), step=5.0,
                                         key=f"pay_amt_{index}")
 
        pay_notes = st.text_input("Payment Notes (optional):", key=f"pay_notes_{index}")
 
        # Show which month is current so user understands the due date logic
        if pay_month == current_month_col:
            st.caption("This is the current month — due date will advance 30 days after payment.")
        else:
            st.caption("This is a past/future month — due date will NOT change.")
 
        if st.button(f"Apply Payment to {dict(MONTHS_DISPLAY).get(pay_month, pay_month)}",
                     key=f"apply_pay_{index}"):
            if pay_amount > 0:
                # Only advance due date if paying the CURRENT month
                is_current_month = (pay_month == current_month_col)
                success, new_due_date = save_payment(
                    customer['Service'], customer['Customer Name'],
                    pay_amount, pay_month=pay_month, notes=pay_notes,
                    advance_due=is_current_month
                )
                if success:
                    if is_current_month and new_due_date:
                        st.success(f"Payment of ${pay_amount:.2f} applied! Due Day updated to {new_due_date.day} (next due: {new_due_date.strftime('%m/%d/%Y')})")
                    elif is_current_month:
                        st.success(f"Payment of ${pay_amount:.2f} applied! Due date advanced.")
                    else:
                        month_name = dict(MONTHS_DISPLAY).get(pay_month, pay_month)
                        st.success(f"Payment of ${pay_amount:.2f} applied to {month_name} (past due — due date unchanged).")
                    st.rerun()
                else:
                    st.error("Error applying payment!")
            else:
                st.warning("Please enter a payment amount greater than $0.")
 

    # ── 4. Monthly Balances ────────────────────────────────
    with st.expander("Monthly Balances (2026)", expanded=False):
        cols = st.columns(6)
        for j, (month_col, month_name) in enumerate(MONTHS_DISPLAY):
            with cols[j % 6]:
                val = monthly_balances.get(month_col, 0)
                label = month_name[:3]
                if val > 0:
                    st.metric(label, f"${val:.2f}")
                else:
                    st.metric(label, "$0.00")
        st.write(f"**Total from months:** ${total_from_aging:.2f}")
 
    # ── 5. Edit Monthly Balance ────────────────────────────
    with st.expander("Edit Monthly Balance"):
        col_edit1, col_edit2 = st.columns(2)
        with col_edit1:
            edit_month = st.selectbox("Select Month:", options=[m[0] for m in MONTHS_DISPLAY],
                                      format_func=lambda x: dict(MONTHS_DISPLAY).get(x, x),
                                      key=f"edit_month_{index}")
        with col_edit2:
            current_val = monthly_balances.get(edit_month, 0)
            new_val = st.number_input("New Balance:", min_value=0.0, value=float(current_val),
                                      step=5.0, key=f"edit_val_{index}")
 
        if st.button(f"Save {dict(MONTHS_DISPLAY).get(edit_month, edit_month)} Balance",
                     key=f"save_{index}"):
            if save_monthly_balance(customer['Service'], customer['Customer Name'], edit_month, new_val):
                update_amount_due_from_months(customer['Service'], customer['Customer Name'])
                st.success("Balance saved!")
                st.rerun()
            else:
                st.error("Error saving balance!")
 

    # ── Square Account (setup / linking) ──────────────────
    with st.expander("🔗 Square Account"):
        if sq_cust_id:
            st.info(f"Customer ID: `{sq_cust_id}`")
            if sq_card_id:
                st.success(f"Card on file: `{sq_card_id}`")
            else:
                st.warning("No card linked yet.")

            col_sq1, col_sq2 = st.columns(2)
            with col_sq1:
                if st.button("List Cards on File", key=f"sq_list_{index}"):
                    cards, err = list_square_cards(sq_cust_id)
                    if err:
                        st.error(f"Error: {err}")
                    elif cards:
                        st.session_state[f'sq_cards_{index}'] = cards
                    else:
                        st.warning("No cards on file in Square.")
            with col_sq2:
                if st.button("Refresh", key=f"sq_refresh_{index}"):
                    st.session_state.pop(f'sq_cards_{index}', None)
                    st.rerun()

            # Show listed cards
            if f'sq_cards_{index}' in st.session_state:
                for card in st.session_state[f'sq_cards_{index}']:
                    cid   = card.get('id', '')
                    brand = card.get('card_brand', '').replace('_', ' ')
                    last4 = card.get('last_4', '???')
                    exp_m = card.get('exp_month', '')
                    exp_y = card.get('exp_year', '')
                    is_active = card.get('enabled', True)
                    label = f"{brand} ...{last4}  exp {exp_m}/{exp_y}" + (" ✅" if is_active else " ❌ disabled")
                    st.write(f"• `{cid}` — {label}")
                    if st.button(f"Use ...{last4} as default", key=f"use_card_{cid}_{index}"):
                        if save_square_ids(customer['Service'], customer['Customer Name'], sq_cust_id, cid):
                            st.success(f"Card ...{last4} saved!")
                            st.session_state.pop(f'sq_cards_{index}', None)
                            st.rerun()
        else:
            st.warning("No Square customer linked yet.")
            if st.button("✨ Create Square Customer", key=f"sq_create_{index}"):
                with st.spinner("Creating customer in Square..."):
                    new_cust_id, err = create_square_customer(
                        customer['Customer Name'], customer.get('Phone', '')
                    )
                if new_cust_id:
                    save_square_ids(customer['Service'], customer['Customer Name'], new_cust_id, '')
                    st.success(f"Customer created! ID: `{new_cust_id}`")
                    st.rerun()
                else:
                    st.error(f"Square error: {err}")

        st.divider()
        st.caption("Or enter IDs manually (get from Square Dashboard → Customers):")
        col_m1, col_m2 = st.columns(2)
        with col_m1:
            manual_cust_id = st.text_input("Square Customer ID", key=f"manual_sq_cust_{index}",
                                           placeholder="e.g. CUST_XXXXXXXX")
        with col_m2:
            manual_card_id = st.text_input("Square Card ID", key=f"manual_sq_card_{index}",
                                           placeholder="e.g. ccof_XXXXXXXX")
        if st.button("💾 Save Square IDs", key=f"save_sq_{index}"):
            if manual_cust_id or manual_card_id:
                cust_to_save = manual_cust_id if manual_cust_id else sq_cust_id
                card_to_save = manual_card_id if manual_card_id else sq_card_id
                if save_square_ids(customer['Service'], customer['Customer Name'],
                                   cust_to_save, card_to_save):
                    st.success("Square IDs saved!")
                    st.rerun()
            else:
                st.warning("Enter at least one ID to save.")

    # Edit customer info
    with st.expander("Edit Customer Info"):
        with st.form(f"edit_form_{index}"):
            new_name = st.text_input("Name", value=str(customer.get('Customer Name', '')))
            phone = st.text_input("Phone", value=str(customer.get('Phone', '')))
            card = st.text_input("Card Number", value=str(customer.get('Card Number', '')))
            exp = st.text_input("Exp", value=str(customer.get('Exp', '')))
            cvv = st.text_input("CVV", value=str(customer.get('CVV', '')))
            plan = st.number_input("Plan Cost", value=safe_float(customer.get('Plan Cost', 0)))
 
            # Modem Numbers
            modem_val = str(customer.get('Modem Numbers', '') or '').strip()
            if modem_val.lower() in ('nan', 'none'):
                modem_val = ''
            modem_input = st.text_input("Modem Number", value=modem_val)
 
            # Notes
            notes_val = str(customer.get('Notes', '') or '')
            notes_input = st.text_area("Notes", value=notes_val)
            notes2_val = str(customer.get('Notes2', '') or '')
            notes2_input = st.text_area("Notes2", value=notes2_val)
 
            submitted = st.form_submit_button("Save All Changes")
            if submitted:
                success = save_customer_info(
                    customer['Service'], customer['Customer Name'],
                    new_name, phone, card, exp, cvv, plan,
                    row_index=customer.get('row_index')
                )
                if success:
                    # Save notes and modem number separately
                    save_customer_notes(customer['Service'], new_name, notes_input)
                    save_notes2(customer['Service'], new_name, notes2_input)
                    save_modem_number(customer['Service'], new_name, modem_input)
                    st.success("Customer info saved!")
                    st.rerun()
                else:
                    st.error("Error saving customer info!")
 
    st.divider()
 
 

def generate_past_due_report(past_due_customers):
    """Generate a printable HTML report of all past due customers with months owed."""
    from datetime import datetime as _dt
    now = _dt.now().strftime('%B %d, %Y %I:%M %p')
    rows = ""
    for c in past_due_customers:
        name     = c.get('Customer Name', '')
        service  = c.get('Service', '')
        phone    = str(c.get('Phone', '') or '')
        card     = mask_card(c.get('Card Number', ''))
        exp      = str(c.get('Exp', '') or '')
        cvv      = str(c.get('CVV', '') or '')
        due_day  = c.get('Due Day', '')
        balance  = get_balance(c)
        modem    = str(c.get('Modem Numbers', '') or '').strip()
        if modem.lower() in ('nan', 'none', ''):
            modem = ''

        # Find which months have a balance
        months_owed = []
        for m in MONTHS_2026:
            val = safe_float(c.get(m, 0))
            if val > 0:
                label = dict(MONTHS_DISPLAY).get(m, m)
                months_owed.append(f"{label}: ${val:.2f}")
        months_str = ', '.join(months_owed) if months_owed else 'Balance on account'

        rows += f"""
        <tr>
            <td>{name}</td>
            <td>{service}</td>
            <td>{phone}</td>
            <td>${balance:.2f}</td>
            <td>{months_str}</td>
            <td>{due_day}</td>
            <td>{card}</td>
            <td>{exp}</td>
            <td>{cvv}</td>
            <td>{modem}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>K-Wireless Past Due Report</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; margin: 20px; color: #000; }}
  h1 {{ font-size: 18px; margin-bottom: 4px; }}
  .meta {{ font-size: 11px; color: #555; margin-bottom: 14px; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th {{ background: #1a1a2e; color: #fff; padding: 6px 8px; text-align: left; font-size: 11px; }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #ddd; vertical-align: top; }}
  tr:nth-child(even) {{ background: #f7f7f7; }}
  .total-row {{ font-weight: bold; background: #eef; }}
  @media print {{
    button {{ display: none; }}
    body {{ margin: 10px; }}
  }}
</style>
</head>
<body>
<button onclick="window.print()" style="margin-bottom:12px;padding:8px 18px;background:#1a1a2e;color:white;border:none;border-radius:4px;cursor:pointer;font-size:13px;">🖨️ Print Report</button>
<h1>K-Wireless — Past Due Customers</h1>
<div class="meta">Generated: {now} &nbsp;|&nbsp; Total past due: <strong>{len(past_due_customers)}</strong> customers</div>
<table>
  <thead>
    <tr>
      <th>Name</th>
      <th>Service</th>
      <th>Phone</th>
      <th>Balance</th>
      <th>Months Owed</th>
      <th>Due Day</th>
      <th>Card</th>
      <th>Exp</th>
      <th>CVV</th>
      <th>Modem #</th>
    </tr>
  </thead>
  <tbody>
    {rows}
    <tr class="total-row">
      <td colspan="3">TOTAL ({len(past_due_customers)} customers)</td>
      <td>${sum(get_balance(c) for c in past_due_customers):.2f}</td>
      <td colspan="6"></td>
    </tr>
  </tbody>
</table>
</body>
</html>"""
    return html

# === MAIN UI ===
 
st.title("📡 K-Wireless Payment Manager")
 
all_data, excel_file = load_excel()
 
if all_data is None:
    st.error("Could not load data. Check that the Excel file exists and is valid.")
    st.stop()
 
# Sidebar
with st.sidebar:
    # ── 1. Daily Collections ───────────────────────────────
    st.header("💰 Collections")

    all_records = parse_collections_from_notes(all_data)
    today = datetime.now().date()

    today_records = [r for r in all_records if r['date'] == today]
    today_total = sum(r['amount'] for r in today_records)
    st.metric("Collected Today", f"${today_total:,.2f}", f"{len(today_records)} payment(s)")

    if today_records:
        for r in today_records:
            st.write(f"• {r['customer']} ({r['service']}) — ${r['amount']:,.2f}")

    st.divider()
    st.subheader("Look Up a Date")
    lookup_date = st.date_input("Select date", value=today, key="collections_date")
    lookup_records = [r for r in all_records if r['date'] == lookup_date]
    lookup_total = sum(r['amount'] for r in lookup_records)

    if lookup_records:
        st.success(f"Total: **${lookup_total:,.2f}** ({len(lookup_records)} payment(s))")
        for r in lookup_records:
            st.write(f"• {r['customer']} ({r['service']}) — ${r['amount']:,.2f}")
    else:
        st.info("No payments recorded on this date.")

    # ── 2. Totals ──────────────────────────────────────────
    st.divider()
    total_customers = 0
    total_outstanding = 0
    service_stats = []
    for service, df in all_data.items():
        if df is not None and not df.empty:
            count = len(df)
            total_customers += count
            revenue = 0
            for _, row in df.iterrows():
                row_total = 0
                for m in MONTHS_2026:
                    row_total += safe_float(row.get(m, 0))
                if row_total == 0:
                    row_total = safe_float(row.get('Amount Due', 0))
                revenue += row_total
            total_outstanding += revenue
            service_stats.append((service, count, revenue))

    st.metric("Total Outstanding", f"${total_outstanding:,.2f}")
    st.metric("Total Customers", total_customers)

    # ── 3. Summary by Service ──────────────────────────────
    st.divider()
    st.header("Summary")
    for service, count, revenue in service_stats:
        st.metric(service, f"{count} customers", f"${revenue:,.2f} outstanding")

    # ── 4. Database Management (bottom) ───────────────────
    st.divider()
    st.subheader("🗄️ Database")
    try:
        with open(EXCEL_FILE, "rb") as db_f:
            st.download_button(
                label="⬇️ Download Database",
                data=db_f,
                file_name="kwireless_billing.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except FileNotFoundError:
        st.warning("Database file not found.")
    uploaded_file = st.file_uploader("⬆️ Upload Database", type=["xlsx"])
    if uploaded_file is not None:
        if uploaded_file.size > 10 * 1024 * 1024:
            st.error("File too large. Maximum 10MB allowed.")
        else:
            try:
                test_wb = load_workbook(uploaded_file)
                sheet_names = test_wb.sheetnames
                if not sheet_names:
                    st.error("Invalid Excel file: no sheets found.")
                else:
                    uploaded_file.seek(0)
                    with open(EXCEL_FILE, "wb") as db_f:
                        db_f.write(uploaded_file.getbuffer())
                    st.success(f"Uploaded {len(sheet_names)} sheet(s). Refresh to reload.")
            except Exception as e:
                st.error(f"Invalid file: {e}")
 
# Main content
st.header("Search & Manage")
 
# Auto-charge section
with st.expander("Auto-Charge Due Today"):
    current_month_name = dict(MONTHS_DISPLAY).get(MONTH_MAP.get(datetime.now().month, ''), '')
    st.write(f"Charges all customers due **today (day {datetime.now().day})** plus any missed in the "
             f"**past 7 days** that haven't been charged or paid yet. "
             f"Catch-up charges are marked with ⚠️.")
    # Show results from previous auto-charge run
    if 'auto_charge_results' in st.session_state:
        results = st.session_state.pop('auto_charge_results')
        if results and not any('Error' in r for r in results):
            st.success(f"Charged {len(results)} customer(s)!")
            for r in results:
                st.write(f"- {r}")
        elif results and any('Error' in r for r in results):
            for r in results:
                st.error(r)
        else:
            st.info("No customers to charge today.")
 
    col_ac1, col_ac2 = st.columns([1, 3])
    with col_ac1:
        if st.button("Run Auto-Charge"):
            with st.spinner("Processing charges..."):
                results = auto_charge_due_today()
                st.session_state['auto_charge_results'] = results
                st.rerun()
 
# Tabs
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "Search", "Filter by Due Date", "Past Due", "Collections History", "Add Customer"
])
 
with tab1:
    search_query = st.text_input("Search by name, phone, or card number:",
                                 placeholder="Enter name or phone...", key="search_name")
    if search_query:
        results = search_customers(all_data, search_query)
        if results:
            st.write(f"Found **{len(results)}** customer(s)")
            for i, customer in enumerate(results):
                with st.container():
                    display_customer_card(customer, i)
        else:
            st.info("No customers found matching your search.")
 
with tab2:
    due_day = st.selectbox("Select Due Day:", options=list(range(1, 32)))
    if due_day:
        customers = get_customers_by_due_day(all_data, due_day)
        st.write(f"Found **{len(customers)}** customers due on day {due_day}")
        if customers:
            # Display as table for quick overview
            table_data = []
            for c in customers:
                modem = str(c.get('Modem Numbers', '') or '').strip()
                if modem.lower() in ('nan', 'none', ''):
                    modem = ''
                table_data.append({
                    'Name': c.get('Customer Name', ''),
                    'Service': c.get('Service', ''),
                    'Balance': f"${safe_float(c.get('Amount Due', 0)):.2f}",
                    'Plan Cost': f"${safe_float(c.get('Plan Cost', 0)):.2f}",
                    'Card': mask_card(c.get('Card Number', '')),
                    'Exp': c.get('Exp', ''),
                    'CVV': c.get('CVV', ''),
                    'Phone': c.get('Phone', ''),
                    'Modem #': modem,
                })
            st.dataframe(pd.DataFrame(table_data), use_container_width=True, hide_index=True)
 
with tab3:
    past_due = get_past_due_customers(all_data)
    total_past_due_bal = sum(get_balance(c) for c in past_due)
    col_pd1, col_pd2 = st.columns([3, 1])
    with col_pd1:
        st.write(f"Found **{len(past_due)}** customers with outstanding balance — Total: **${total_past_due_bal:,.2f}**")
    with col_pd2:
        if past_due:
            report_html = generate_past_due_report(past_due)
            st.download_button(
                label="🖨️ Print / Download Report",
                data=report_html,
                file_name="past_due_report.html",
                mime="text/html",
                use_container_width=True
            )
    if past_due:
        table_data = []
        for c in past_due:
            balance = get_balance(c)
            modem = str(c.get('Modem Numbers', '') or '').strip()
            if modem.lower() in ('nan', 'none', ''):
                modem = ''
            # Which months are owed
            months_owed = []
            for m in MONTHS_2026:
                val = safe_float(c.get(m, 0))
                if val > 0:
                    months_owed.append(f"{dict(MONTHS_DISPLAY).get(m, m)}: ${val:.2f}")
            table_data.append({
                'Name': c.get('Customer Name', ''),
                'Service': c.get('Service', ''),
                'Phone': c.get('Phone', ''),
                'Balance': f"${balance:.2f}",
                'Months Owed': ', '.join(months_owed) if months_owed else '',
                'Due Day': c.get('Due Day', 'N/A'),
                'Card': mask_card(c.get('Card Number', '')),
                'Exp': c.get('Exp', ''),
                'CVV': c.get('CVV', ''),
                'Modem #': modem,
            })
        st.dataframe(pd.DataFrame(table_data), use_container_width=True, hide_index=True)
 
with tab4:
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Start Date")
    with col2:
        end_date = st.date_input("End Date")
    if start_date and end_date:
        collections = get_collections_by_date(all_data, start_date, end_date)
        st.write(f"Found **{len(collections)}** payments")
        if collections:
            st.dataframe(pd.DataFrame(collections), use_container_width=True, hide_index=True)
 
with tab5:
    st.subheader("Add New Customer")
    with st.form("add_customer_form"):
        add_service = st.selectbox("Service:", options=list(all_data.keys()))
        add_name = st.text_input("Customer Name:")
        add_phone = st.text_input("Phone:")
        add_card = st.text_input("Card Number:")
        add_exp = st.text_input("Expiration:")
        add_cvv = st.text_input("CVV:")
        add_plan = st.number_input("Plan Cost:", min_value=0.0, step=5.0)
        add_due = st.number_input("Due Day (1-28):", min_value=1, max_value=28, value=1)
 
        if st.form_submit_button("Add Customer"):
            if add_name and add_service:
                try:
                    wb = load_workbook(EXCEL_FILE)
                    ws = wb[add_service]
                    headers = get_header_map(ws)
                    new_row = ws.max_row + 1
 
                    col_vals = {
                        'Charge Date': datetime.now().strftime('%Y-%m-%d'),
                        'Service': add_service,
                        'Plan Cost': add_plan,
                        'Customer Name': add_name,
                        'Card Number': add_card,
                        'Exp': add_exp,
                        'CVV': add_cvv,
                        'Amount Due': 0,
                        'Status': '',
                        'Phone': add_phone,
                        'Due Day': add_due,
                    }
                    for col_name, val in col_vals.items():
                        col_idx = headers.get(col_name)
                        if col_idx:
                            ws.cell(row=new_row, column=col_idx, value=val)
 
                    locked_save(wb, EXCEL_FILE)
                    st.success(f"Added {add_name} to {add_service}!")
                    logger.info(f"Added new customer: {add_name} to {add_service}")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error adding customer: {e}")
                    logger.error(f"Error adding customer: {e}")
            else:
                st.warning("Please enter at least a name and select a service.")
 
# ── COLLECTIONS REPORT ─────────────────────────────────────────────────────
st.divider()
st.subheader("📊 Collections Report")

_report_range = st.radio(
    "Report Period:",
    ["Today", "This Week", "Custom Range"],
    horizontal=True,
    key="report_range_selector"
)

_today = datetime.now().date()
if _report_range == "Today":
    _r_start, _r_end = _today, _today
elif _report_range == "This Week":
    from datetime import timedelta as _td
    _week_start = _today - __import__('datetime').timedelta(days=_today.weekday())
    _r_start, _r_end = _week_start, _today
else:
    _col1, _col2 = st.columns(2)
    _r_start = _col1.date_input("From:", value=_today, key="report_from")
    _r_end   = _col2.date_input("To:",   value=_today, key="report_to")

if st.button("🔍 Generate Report", key="gen_collections_report"):
    with st.spinner("Scanning payment records..."):
        _entries, _err = get_collections_report(_r_start, _r_end)
    if _err:
        st.error(f"Error loading data: {_err}")
    elif not _entries:
        st.info("No payments found for this period.")
    else:
        _total        = sum(e['amount'] for e in _entries)
        _sq_total     = sum(e['amount'] for e in _entries if e['method'] == 'Square')
        _manual_total = sum(e['amount'] for e in _entries if e['method'] == 'Manual')

        # Summary metrics
        _mc1, _mc2, _mc3 = st.columns(3)
        _mc1.metric("Total Collected",    f"${_total:.2f}")
        _mc2.metric("Square Payments",    f"${_sq_total:.2f}")
        _mc3.metric("Manual Payments",    f"${_manual_total:.2f}")

        # Table
        _df = pd.DataFrame([{
            'Customer':  e['customer'],
            'Service':   e['sheet'],
            'Amount':    f"${e['amount']:.2f}",
            'Month':     e['month'],
            'Method':    e['method'],
            'Date/Time': e['dt'].strftime('%m/%d/%Y %I:%M %p'),
        } for e in _entries])
        st.dataframe(_df, use_container_width=True, hide_index=True)
        st.caption(f"{len(_entries)} transaction(s)")

        # Download printable report
        if _r_start == _r_end:
            _period_label = _r_start.strftime('%m/%d/%Y')
        else:
            _period_label = f"{_r_start.strftime('%m/%d/%Y')} \u2013 {_r_end.strftime('%m/%d/%Y')}"
        _html = _build_report_html(_entries, _period_label, _total, _sq_total, _manual_total)
        _fname = f"kwireless_report_{_r_start.strftime('%Y%m%d')}"
        if _r_start != _r_end:
            _fname += f"_{_r_end.strftime('%Y%m%d')}"
        _fname += ".html"
        st.download_button(
            label="📄 Download Printable Report",
            data=_html,
            file_name=_fname,
            mime="text/html",
            help="Download and open in your browser, then use File → Print (or Ctrl+P) to print or save as PDF"
        )

# Footer
st.divider()
st.caption("K-Wireless Payment Manager v2.0")
 
 
 
 
 
